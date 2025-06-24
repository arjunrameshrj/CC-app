import streamlit as st
import pandas as pd
import os
import io
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import plotly.express as px
import re

# --- Configuration ---
st.set_page_config(page_title="KPI Dashboard", layout="wide")

# --- CSS Styling ---
def load_css():
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;600;700&display=swap');

    html, body, [data-testid="stAppViewContainer"] {
        background: linear-gradient(-45deg, #1f4037, #99f2c8, #2c3e50, #3498db);
        background-size: 400% 400%;
        animation: gradient 20s ease infinite;
        color: #ffffff;
        font-family: 'Inter', sans-serif;
    }
    @keyframes gradient {
        0% {background-position: 0% 50%;}
        50% {background-position: 100% 50%;}
        100% {background-position: 0% 50%;}
    }

    .main-container {
        background: rgba(255, 255, 255, 0.12);
        backdrop-filter: blur(10px);
        border-radius: 20px;
        padding: 25px;
        margin-bottom: 30px;
    }

    h1, h2, h3, h4, h5, h6 {
        color: white !important;
        text-shadow: 1px 1px 3px rgba(0,0,0,0.4);
    }

    .stDataFrame {
        border-radius: 15px !important;
        overflow: hidden !important;
        box-shadow: 0 8px 32px rgba(0, 0, 0, 0.3) !important;
        border: 1px solid rgba(255, 255, 255, 0.2) !important;
        background: rgba(255, 255, 255, 0.1) !important;
        backdrop-filter: blur(8px) !important;
    }

    div.stButton > button, div.stDownloadButton > button {
        background: linear-gradient(135deg, #f7971e, #ffd200) !important;
        color: black !important;
        padding: 12px 25px !important;
        border-radius: 50px !important;
        font-weight: 600 !important;
        cursor: pointer !important;
        transition: all 0.3s ease !important;
        border: none !important;
        box-shadow: 0 4px 8px rgba(0, 0, 0, 0.2) !important;
    }

    .kpi-card {
        background: rgba(255, 255, 255, 0.15) !important;
        border-radius: 15px !important;
        padding: 20px !important;
        margin-bottom: 15px !important;
        box-shadow: 0 6px 12px rgba(0, 0, 0, 0.2) !important;
        transition: all 0.3s ease !important;
        border-left: 6px solid;
        position: relative;
        overflow: hidden;
        backdrop-filter: blur(8px);
    }

    .kpi-exceeds {
        border-left-color: #00E676 !important;
        background: linear-gradient(135deg, rgba(0, 230, 118, 0.15), rgba(0, 200, 83, 0.2)) !important;
    }
    .kpi-below {
        border-left-color: #FF5252 !important;
        background: linear-gradient(135deg, rgba(255, 82, 82, 0.15), rgba(255, 41, 41, 0.2)) !important;
    }
    .kpi-neutral {
        border-left-color: #FFD600 !important;
        background: linear-gradient(135deg, rgba(255, 214, 0, 0.15), rgba(255, 171, 0, 0.2)) !important;
    }
    </style>
    """, unsafe_allow_html=True)

# --- Helper Functions ---
def extract_numeric(text):
    """Extracts numeric value from string, handling percentages and time formats."""
    if pd.isna(text):
        return 0.0
    if isinstance(text, (int, float)):
        return float(text)
    
    s = str(text).replace(',', '').strip()

    # Handle time formats (e.g., "2 minutes 30 seconds")
    if "minute" in s or "second" in s:
        minutes_match = re.search(r'(\d+)\s*minute', s)
        seconds_match = re.search(r'(\d+)\s*second', s)
        total_seconds = 0
        if minutes_match:
            total_seconds += int(minutes_match.group(1)) * 60
        if seconds_match:
            total_seconds += int(seconds_match.group(1))
        return float(total_seconds)    
    
    # Handle percentages
    percent_match = re.search(r'(\d+\.?\d*)\%', s)
    if percent_match:
        return float(percent_match.group(1))
    
    # Handle other numeric formats
    num_match = re.search(r'([-+]?\d+\.?\d*)', s)
    if num_match:
        return float(num_match.group(1))
    
    return 0.0

def validate_input(text, column_type):
    """Validates input for Target, Actuals, and Variance columns."""
    if pd.isna(text) or text == "":
        return False, "Value cannot be empty."
    
    text = str(text).strip()
    
    if column_type in ["Target", "Actuals"]:
        # Allow time formats (e.g., "2 minutes 30 seconds") or percentages (e.g., "80%")
        if "minute" in text or "second" in text:
            time_match = re.search(r'(\d+\s*minute[s]?)?\s*(\d+\s*second[s]?)?', text)
            if time_match:
                return True, ""
            return False, "Invalid time format. Use 'X minutes Y seconds'."
        if "%" in text:
            percent_match = re.search(r'(\d+\.?\d*)\%', text)
            if percent_match:
                return True, ""
            return False, "Invalid percentage format. Use 'XX%'."
        num_match = re.search(r'[-+]?\d+\.?\d*', text)
        if num_match:
            return True, ""
        return False, "Invalid numeric format."
    
    if column_type == "Variance":
        # Allow formats like "+5%", "-30 seconds", or numeric values
        if "%" in text or "second" in text or "minute" in text:
            return True, ""
        num_match = re.search(r'[-+]?\d+\.?\d*', text)
        if num_match:
            return True, ""
        return False, "Invalid variance format. Use 'X%', 'X seconds', or numeric value."
    
    return True, ""

def get_kpi_type(kpi_name):
    """Determines if KPI is 'lower_is_better' or 'higher_is_better'."""
    lower_is_better_kpis = [
        "Average Handle Time (AHT)", "Call Abandonment Rate", "High Risk Escalation Rate"
    ]
    for kpi_part in lower_is_better_kpis:
        if kpi_part in kpi_name:
            return 'lower_is_better'
    return 'higher_is_better'

def get_performance_status(actual, target, kpi_type):
    """Compares actual to target and returns performance status."""
    if pd.isna(actual) or pd.isna(target):
        return 'neutral'
        
    if kpi_type == 'lower_is_better':
        return 'exceeds' if actual <= target else 'below'
    else:
        return 'exceeds' if actual >= target else 'below'

def create_status_badge(status):
    """Creates plain text status for dataframe display."""
    return 'Exceeds' if status == 'exceeds' else 'Below' if status == 'below' else 'Neutral'

def initialize_data():
    """Initialize or load KPI data with default values for new rows."""
    default_data = {
        "KPI": [
            "Average Handle Time (AHT)",
            "First Call Resolution (FCR) (48hrs)",
            "Customer Satisfaction Score (CSAT)",
            "Service Level (SL)",
            "Call Abandonment Rate",
            "Call Quality Score",
            "Training Completion Rate",
            "Attendance Rate",
            "Compliance Score",
            "High Risk Escalation Rate"
        ],
        "Target": ["2 minutes", "80%", "90%", "90% (20 secs)", "<7%", "90%", "100%", "95%", "95%", "<5%"],
        "Actuals": ["1 minute 30 seconds", "78%", "88%", "85% (20 secs)", "3%", "80%", "90%", "93%", "92%", "6%"],
        "Variance": ["-30 seconds", "-2%", "-2%", "-5%", "-4%", "-10%", "-10%", "-2%", "-3%", "+1%"],
        "Comments": [
            "Good performance; reduced handle time effectively.",
            "Slightly below target; review training needs.",
            "Consistent but needs improvement in certain areas.",
            "Below target; improve answer speed.",
            "Positive outcome; effective wait management.",
            "Needs attention; consider team-building activities.",
            "Schedule follow-up sessions for new hires.",
            "Monitor patterns; address any absenteeism issues.",
            "Review compliance training; address knowledge gaps.",
            "Slightly above target; investigate root cause."
        ],
        "Explanation of KPIs": [
            "Measures average call duration. Lower is better if quality is maintained.",
            "Calls resolved on first contact. Higher % is better.",
            "Customer feedback rating. Higher % indicates satisfaction.",
            "Calls answered within 20 secs. Higher % is better.",
            "Lower % means fewer abandoned calls.",
            "Quality of agent performance. Higher is better.",
            "Completion of mandatory training. Should be 100%.",
            "Employee attendance reliability. Higher % is better.",
            "Regulatory adherence. Higher % avoids penalties.",
            "Escalation to management. Lower % is ideal."
        ]
    }
    
    DATA_FILE = "kpi_data.csv"
    if os.path.exists(DATA_FILE):
        return pd.read_csv(DATA_FILE)
    return pd.DataFrame(default_data)

def to_excel(df_to_export):
    """Convert DataFrame to styled Excel file."""
    output = io.BytesIO()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "KPI Dashboard"
    
    # Add title
    sheet.merge_cells('A1:F1')
    title_cell = sheet['A1']
    title_cell.value = "KPI Performance Dashboard"
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add headers
    headers = list(df_to_export.columns)
    sheet.append(headers)
    
    # Style headers
    for col_idx in range(1, len(headers) + 1):
        header_cell = sheet.cell(row=2, column=col_idx)
        header_cell.font = Font(bold=True, color="FFFFFF")
        header_cell.fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
        header_cell.alignment = Alignment(horizontal='center')
    
    # Add data and style variance column
    variance_col_idx = headers.index("Variance") + 1 if "Variance" in headers else -1
    kpi_col_idx = headers.index("KPI") + 1 if "KPI" in headers else -1
    
    for r_idx, row_data in enumerate(df_to_export.itertuples(index=False), start=3):
        sheet.append(row_data)
        
        for col_idx in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=r_idx, column=col_idx)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            if col_idx == variance_col_idx and variance_col_idx != -1:
                kpi_name = sheet.cell(row=r_idx, column=kpi_col_idx).value
                variance_str = str(cell.value)
                variance_num = extract_numeric(variance_str)
                kpi_type = get_kpi_type(kpi_name)

                if kpi_type == 'lower_is_better':
                    cell.font = Font(color="00E676" if variance_num < 0 else "FF5252" if variance_num > 0 else "FFD600", bold=True)
                else:
                    cell.font = Font(color="00E676" if variance_num > 0 else "FF5252" if variance_num < 0 else "FFD600", bold=True)

    # Adjust column widths
    for col_idx in range(1, sheet.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in sheet[column_letter]:
            try:
                if cell.value is not None and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Add borders
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=sheet.max_column):
        for cell in row:
            cell.border = thin_border
    
    workbook.save(output)
    return output.getvalue()

# --- Updated Editor Function ---
def display_editor(df):
    """Display the KPI editor view with table editing, row addition, and deletion functionality."""
    st.subheader("📝 Edit KPI Metrics")
    
    # Instructions for users
    st.markdown("""
    **Instructions:**
    - **Edit**: Click on any cell to edit its value.
    - **Add Row**: Click the ➕ button at the bottom of the table to add a new KPI.
    - **Delete Row**: Select rows using the checkbox in the '🗑️ Select to Delete' column and click the '🗑️ Delete Selected Rows' button.
    - **Validate**: Ensure 'Target', 'Actuals', and 'Variance' are in valid formats (e.g., '80%', '2 minutes 30 seconds', or numeric values).
    - **Save**: Click 'Save Changes' to save your edits to the CSV file.
    - **Download**: Export the table as an Excel file using 'Download Excel'.
    """, unsafe_allow_html=True)
    
    # Initialize session state for edited dataframe
    if 'edited_df' not in st.session_state:
        st.session_state.edited_df = df.copy()
    
    # Add a temporary column for selection before displaying in data_editor
    df_for_editor = st.session_state.edited_df.copy()
    df_for_editor.insert(0, 'Select', False) # Add a boolean column for selection at the beginning

    # Column configuration with validation
    column_config = {
        "Select": st.column_config.CheckboxColumn(
            "🗑️ Select to Delete",
            help="Select rows to delete",
            default=False,
            width="small"
        ),
        "KPI": st.column_config.TextColumn(
            "KPI Name",
            help="Enter the name of the KPI (e.g., 'Customer Satisfaction Score (CSAT)')",
            required=True
        ),
        "Target": st.column_config.TextColumn(
            "Target Value",
            help="Enter target value (e.g., '80%', '2 minutes', or numeric)",
            required=True
        ),
        "Actuals": st.column_config.TextColumn(
            "Current Value",
            help="Enter actual value (e.g., '78%', '1 minute 30 seconds', or numeric)",
            required=True
        ),
        "Variance": st.column_config.TextColumn(
            "Variance",
            help="Enter variance (e.g., '-2%', '-30 seconds', or numeric)",
            required=True
        ),
        "Comments": st.column_config.TextColumn(
            "Comments",
            help="Additional notes about this KPI",
            default=""
        ),
        "Explanation of KPIs": st.column_config.TextColumn(
            "KPI Explanation",
            help="Description of what this KPI measures",
            default=""
        )
    }
    
    # Display editable dataframe with row selection and dynamic rows
    edited_df_result = st.data_editor(
        df_for_editor, # Pass the DataFrame with the 'Select' column
        num_rows="dynamic",
        use_container_width=True,
        column_config=column_config,
        key="kpi_data_editor",
        hide_index=True,
        disabled=False
    )
    
    # Update the session state with the result from data_editor, removing the 'Select' column
    st.session_state.edited_df = edited_df_result.drop(columns=['Select'], errors='ignore')

    # Validate edited data (excluding the 'Select' column for validation)
    validation_errors = []
    for idx, row in edited_df_result.drop(columns=['Select'], errors='ignore').iterrows():
        for col in ["Target", "Actuals", "Variance"]:
            is_valid, error_msg = validate_input(row[col], col)
            if not is_valid:
                validation_errors.append(f"Row {idx + 1}, {col}: {error_msg}")
    
    if validation_errors:
        st.error("🚨 Validation Errors:\n" + "\n".join(validation_errors))
    
    # Action buttons
    col1, col2, col3 = st.columns([1, 1, 1]) # Added a column for the delete button
    
    with col1:
        if st.button("💾 Save Changes"):
            if not validation_errors:
                # Save the DataFrame without the 'Select' column
                st.session_state.current_df = st.session_state.edited_df
                st.session_state.current_df.to_csv("kpi_data.csv", index=False)
                st.success("✅ Changes saved successfully!")
                st.rerun()  # Updated to st.rerun() for Streamlit >= 1.10
            else:
                st.error("🚨 Please fix validation errors before saving.")
    
    with col2:
        if st.button("🗑️ Delete Selected Rows"):
            # Filter out selected rows
            rows_to_keep = edited_df_result[edited_df_result['Select'] == False].drop(columns=['Select'])
            if len(rows_to_keep) < len(edited_df_result.drop(columns=['Select'])):
                st.session_state.current_df = rows_to_keep
                st.session_state.edited_df = rows_to_keep.copy() # Update edited_df as well
                st.session_state.current_df.to_csv("kpi_data.csv", index=False)
                st.success("🗑️ Selected rows deleted successfully!")
                st.rerun()
            else:
                st.warning("No rows selected for deletion.")

    with col3:
        excel_data = to_excel(st.session_state.edited_df) # Use the current edited_df for export
        st.download_button(
            label="⬇️ Download Excel",
            data=excel_data,
            file_name="kpi_dashboard.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# --- Main App ---
def main():
    load_css()
    
    # Initialize data
    if 'current_df' not in st.session_state:
        st.session_state.current_df = initialize_data()
    
    df = st.session_state.current_df
    
    # Banner and Title
    st.markdown("<h1>📊 KPI Performance Dashboard</h1>", unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)
    
    # Create tabs
    tab1, tab2 = st.tabs(["Dashboard", "Edit KPIs"])

    with tab1:
        # Dashboard View
        display_dashboard(df)
    
    with tab2:
        # KPI Editor View
        display_editor(df)

def display_dashboard(df):
    """Display the dashboard view with KPIs and visualizations."""
    # Enhanced KPI Summary Cards
    st.subheader("📊 KPI Summary")
    
    card_data = df.copy()
    card_data["Actual_Numeric"] = card_data["Actuals"].apply(extract_numeric)
    card_data["Target_Numeric"] = card_data["Target"].apply(extract_numeric)
    card_data["KPI_Type"] = card_data["KPI"].apply(get_kpi_type)
    card_data["Performance_Status"] = card_data.apply(
        lambda row: get_performance_status(row["Actual_Numeric"], row["Target_Numeric"], row["KPI_Type"]),
        axis=1
    )
    
    # Create columns for cards - 5 per row
    num_kpis = len(card_data)
    cols_per_row = 5
    num_rows = (num_kpis + cols_per_row - 1) // cols_per_row

    for i in range(num_rows):
        cols = st.columns(cols_per_row)
        for idx in range(i * cols_per_row, min((i + 1) * cols_per_row, num_kpis)):
            row = card_data.iloc[idx]
            status = row['Performance_Status']
            
            # Determine icon and status class
            if status == 'exceeds':
                icon = '🚀'
                status_class = 'kpi-exceeds'
                badge_color = "linear-gradient(135deg, #00E676, #00C853)"
            elif status == 'below':
                icon = '⚠️'
                status_class = 'kpi-below'
                badge_color = "linear-gradient(135deg, #FF5252, #FF1744)"
            else:
                icon = '🔍'
                status_class = 'kpi-neutral'
                badge_color = "linear-gradient(135deg, #FFD600, #FFAB00)"
            
            with cols[idx % cols_per_row]:
                st.markdown(
                    f"""
                    <div class="kpi-card {status_class}">
                        <h3><span class="kpi-icon">{icon}</span> {row['KPI'].split('(')[0].strip()}</h3>
                        <div class="value">{row['Actuals']}</div>
                        <div class="target">Target: {row['Target']}</div>
                        <div class="variance">Variance: {row['Variance']}</div>
                        <div style="margin-top: 10px; background: {badge_color}; padding: 5px 10px; border-radius: 50px; display: inline-block; font-size: 12px; font-weight: 600; color: white; text-shadow: 1px 1px 2px rgba(0,0,0,0.2);">
                            {icon} {'Exceeds' if status == 'exceeds' else 'Below' if status == 'below' else 'Neutral'}
                        </div>
                    </div>
                    """,
                    unsafe_allow_html=True
                )

    # Current KPI Performance - Table View
    st.subheader("📈 Current KPI Performance")
    
    display_df = df.copy()
    display_df["Actual_Numeric"] = display_df["Actuals"].apply(extract_numeric)
    display_df["Target_Numeric"] = display_df["Target"].apply(extract_numeric)
    display_df["KPI_Type"] = display_df["KPI"].apply(get_kpi_type)
    display_df["Performance_Status"] = display_df.apply(
        lambda row: get_performance_status(row["Actual_Numeric"], row["Target_Numeric"], row["KPI_Type"]),
        axis=1
    )
    
    table_df = display_df[["KPI", "Target", "Actuals", "Variance", "Comments"]].copy()
    table_df["Status"] = display_df["Performance_Status"].apply(create_status_badge)
    
    def style_row(row):
        variance_str = str(row['Variance'])
        kpi_type = get_kpi_type(row['KPI'])
        variance_num = extract_numeric(variance_str)
        
        if kpi_type == 'lower_is_better':
            variance_color = '#00E676' if variance_num < 0 else '#FF5252' if variance_num > 0 else '#FFD600'
        else:
            variance_color = '#00E676' if variance_num > 0 else '#FF5252' if variance_num < 0 else '#FFD600'
        
        status = row['Status']
        status_color = '#00E676' if status == 'Exceeds' else '#FF5252' if status == 'Below' else '#FFD600'
        
        return [
            '' if col != 'Variance' and col != 'Status' else
            f'color: {variance_color}; font-weight: bold' if col == 'Variance' else
            f'color: {status_color}; font-weight: bold; background: {status_color}20; border-radius: 10px; padding: 5px;'
            for col in table_df.columns
        ]
    
    styled_df = table_df.style.apply(style_row, axis=1).set_properties(**{
        'text-align': 'center',
        'font-size': '14px',
        'padding': '12px',
        'border': 'none'
    })
    
    st.dataframe(styled_df, use_container_width=True, hide_index=True)

    # Enhanced Visualizations
    st.subheader("📊 Key Performance Insights")
    
    try:
        visual_df = df.copy()
        visual_df["Actual_Numeric"] = visual_df["Actuals"].apply(extract_numeric)
        visual_df["Target_Numeric"] = visual_df["Target"].apply(extract_numeric)
        visual_df["KPI_Type"] = visual_df["KPI"].apply(get_kpi_type)
        visual_df["Performance_Status"] = visual_df.apply(
            lambda row: get_performance_status(row["Actual_Numeric"], row["Target_Numeric"], row["KPI_Type"]),
            axis=1
        )
        
        viz_tab1, viz_tab2 = st.tabs(["Performance Overview", "Target Achievement"])
        
        with viz_tab1:
            fig1 = px.bar(
                visual_df,
                x="KPI",
                y=["Target_Numeric", "Actual_Numeric"],
                barmode="group",
                title="<b>Actual vs Target Performance</b>",
                color_discrete_map={"Target_Numeric": "#4F81BD", "Actual_Numeric": "#FFA500"},
                height=500
            )
            fig1.update_layout(
                xaxis_tickangle=-45,
                plot_bgcolor='rgba(0,0,0,0)',
                paper_bgcolor='rgba(0,0,0,0)',
                font=dict(color="white")
            )
            st.plotly_chart(fig1, use_container_width=True)
            
        with viz_tab2:
            performance_counts = visual_df["Performance_Status"].value_counts().reset_index()
            performance_counts.columns = ["Status", "Count"]
            
            fig2 = px.pie(
                performance_counts,
                values="Count",
                names="Status",
                title="<b>Overall Target Achievement</b>",
                color="Status",
                color_discrete_map={"exceeds": "#00E676", "below": "#FF5252", "neutral": "#FFD600"},
                hole=0.4,
                height=500
            )
            fig2.update_layout(
                plot_bgcolor='rgba(0,0,0,0)',
                paper_bgcolor='rgba(0,0,0,0)',
                font=dict(color="white")
            )
            st.plotly_chart(fig2, use_container_width=True)
            
    except Exception as e:
        st.warning(f"📊 Some charts could not be rendered. Error: {str(e)}")

    # Performance Summary
    st.subheader("📋 Performance Summary")
    status_df = df.copy()
    status_df["Actual_Numeric"] = status_df["Actuals"].apply(extract_numeric)
    status_df["Target_Numeric"] = status_df["Target"].apply(extract_numeric)
    status_df["KPI_Type"] = status_df["KPI"].apply(get_kpi_type)
    status_df["Performance_Status"] = status_df.apply(
        lambda row: get_performance_status(row["Actual_Numeric"], row["Target_Numeric"], row["KPI_Type"]),
        axis=1
    )
    
    status_display_map = {
        'exceeds': "✅ Exceeds Target",
        'below': "⚠️ Below Target",
        'neutral': "🔍 Neutral"
    }
    status_df["Display_Status"] = status_df["Performance_Status"].map(status_display_map)

    cols = st.columns(5)
    metrics = status_df["Display_Status"].value_counts()
    
    with cols[0]:
        st.metric("✅ Exceeds Target", metrics.get("✅ Exceeds Target", 0))
    with cols[1]:
        st.metric("⚠️ Below Target", metrics.get("⚠️ Below Target", 0))
    with cols[2]:
        total_kpis = len(status_df)
        st.metric("📊 Total KPIs Tracked", total_kpis)
    with cols[3]:
        percent_achieved = round((metrics.get("✅ Exceeds Target", 0) / total_kpis) * 100 if total_kpis > 0 else 0, 1)
        st.metric("🎯 Target Achievement", f"{percent_achieved}%")

if __name__ == "__main__":
    main()
